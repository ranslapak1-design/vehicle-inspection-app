import openpyxl

path = r"C:\Users\Ran Slapak\Desktop\יצרנים\וולוו\14.2.2026\ר.ר-וולוו 481084 SO26L00762\ר.ר-וולוו 481084 SO26L00762.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)

print("=" * 80)
print("PART 1: Searching 'בוחן' sheet for cells containing 'חוסרים' or 'הערות'")
print("=" * 80)

ws = wb["בוחן"]
for row in range(1, 501):
    for col in range(1, 30):
        cell = ws.cell(row=row, column=col)
        if cell.value and isinstance(cell.value, str):
            if "חוסרים" in cell.value or "הערות" in cell.value:
                print(f"  {cell.coordinate}: {repr(cell.value)}")

print()
print("=" * 80)
print("PART 2: ALL non-empty cells in 'פ. ממצאים מסכם' rows 1-50")
print("=" * 80)

ws2 = wb["פ. ממצאים מסכם"]
for row in range(1, 51):
    for col in range(1, 30):
        cell = ws2.cell(row=row, column=col)
        if cell.value is not None:
            print(f"  {cell.coordinate}: {repr(cell.value)}")

print()
print("=" * 80)
print("PART 3: Columns B and H for rows 20-40 in 'פ. ממצאים מסכם'")
print("=" * 80)

for row in range(20, 41):
    b = ws2.cell(row=row, column=2)
    h = ws2.cell(row=row, column=8)
    print(f"  Row {row}: B={repr(b.value)}  |  H={repr(h.value)}")

wb.close()
