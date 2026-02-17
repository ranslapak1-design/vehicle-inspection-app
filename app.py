# -*- coding: utf-8 -*-
import os
import json
import base64
from datetime import datetime
from pathlib import Path

from flask import Flask, render_template, request, jsonify, send_file

import openpyxl
from fpdf import FPDF
import io
import tempfile

BASE_DIR = Path(r"C:\Users\Ran Slapak\Desktop\יצרנים")
APP_DIR = Path(__file__).parent.resolve()

app = Flask(__name__, template_folder=str(APP_DIR / "templates"))
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB

# ---------------------------------------------------------------------------
# Directory helpers
# ---------------------------------------------------------------------------

def list_manufacturers():
    """List manufacturer folders in BASE_DIR."""
    if not BASE_DIR.is_dir():
        return []
    return sorted([d.name for d in BASE_DIR.iterdir() if d.is_dir()])


def list_dates(manufacturer):
    """List date folders under a manufacturer."""
    mfr_dir = BASE_DIR / manufacturer
    if not mfr_dir.is_dir():
        return []
    return sorted([d.name for d in mfr_dir.iterdir() if d.is_dir()], reverse=True)


def list_vehicles(manufacturer, date_folder):
    """List vehicle folders under a date."""
    date_dir = BASE_DIR / manufacturer / date_folder
    if not date_dir.is_dir():
        return []
    vehicles = []
    for d in sorted(date_dir.iterdir()):
        if d.is_dir():
            # Check if matching xlsx exists inside
            xlsx = d / f"{d.name}.xlsx"
            vehicles.append({
                "name": d.name,
                "has_excel": xlsx.is_file(),
            })
    return vehicles


def get_vehicle_path(manufacturer, date_folder, vehicle):
    """Get full path to a vehicle folder."""
    return BASE_DIR / manufacturer / date_folder / vehicle


def get_excel_path(manufacturer, date_folder, vehicle):
    """Get full path to the vehicle's Excel file."""
    vdir = get_vehicle_path(manufacturer, date_folder, vehicle)
    return vdir / f"{vehicle}.xlsx"


def get_photos_dir(manufacturer, date_folder, vehicle):
    """Get the photos directory for a vehicle."""
    return get_vehicle_path(manufacturer, date_folder, vehicle) / "תמונות"


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------

# Cell mapping: מזכירה reference data (source of truth)
SECRETARY_CELLS = {
    "N": {  # N2/N3
        "license":       ("מזכירה", "D16"),
        "category":      ("מזכירה", "D17"),
        "tire_front":    ("מזכירה", "D22"),
        "tire_rear":     ("מזכירה", "D23"),
        "total_weight":  ("מזכירה", "D24"),
        "vin":           ("מזכירה", "D25"),
        "num_wheels":    ("מזכירה", "D29"),
        "color":         ("מזכירה", "D30"),
        "axle_distance": ("מזכירה", "D32"),
        "manufacturer":  ("מזכירה", "D33"),
        "num_axles":     ("מזכירה", "D42"),
        "tire_front_hr": ("מזכירה", "D44"),
        "tire_rear_hr":  ("מזכירה", "D47"),
        "wvta":          ("מזכירה", "D50"),
        "weight_total":  ("מזכירה", "D51"),
        "weight_front":  ("מזכירה", "D52"),
        "weight_rear":   ("מזכירה", "D53"),
        "axle_dist_hr":  ("מזכירה", "D55"),
        # Device 1 secretary reference
        "sec_dev1_name":         ("מזכירה", "D102"),
        "sec_dev1_installer":    ("מזכירה", "D100"),
        "sec_dev1_manufacturer": ("מזכירה", "D108"),
        "sec_dev1_model":        ("מזכירה", "D109"),
        "sec_dev1_serial":       ("מזכירה", "D110"),
        # Device 2 secretary reference
        "sec_dev2_name":         ("מזכירה", "D128"),
        "sec_dev2_installer":    ("מזכירה", "D126"),
        "sec_dev2_manufacturer": ("מזכירה", "D141"),
        "sec_dev2_model":        ("מזכירה", "D142"),
        "sec_dev2_serial":       ("מזכירה", "D143"),
        # Device 3 secretary reference
        "sec_dev3_name":         ("מזכירה", "D154"),
        "sec_dev3_installer":    ("מזכירה", "D152"),
        "sec_dev3_manufacturer": ("מזכירה", "D167"),
        "sec_dev3_model":        ("מזכירה", "D168"),
        "sec_dev3_serial":       ("מזכירה", "D169"),
    },
    "M": {  # M1
        "license":       ("מזכירה", "D16"),
        "category":      ("מזכירה", "D17"),
        "tire_front":    ("מזכירה", "D22"),
        "tire_rear":     ("מזכירה", "D23"),
        "total_weight":  ("מזכירה", "D24"),
        "vin":           ("מזכירה", "D25"),
        "color":         ("מזכירה", "D30"),
        "axle_distance": ("מזכירה", "D32"),
        "manufacturer":  ("מזכירה", "D33"),
    },
}

# Cell mapping: בוחן sheet (where examiner writes)
EXAMINER_CELLS_N = {
    "license":        "E42",
    "color":          "E43",
    "seats_beside":   "E44",
    "seats_behind":   "E45",
    "sleeping":       "E46",
    "num_axles":      "E48",
    "num_wheels":     "E49",
    "tire1":          "E50",
    "tire2":          "E51",
    "tire3":          "E52",
    "tire4":          "E53",
    "vin":            "E63",
    "category":       "E65",
    "manufacturer":   "E69",
    "wvta":           "E70",
    "weight_total":   "E72",
    "weight_coupled": "E73",
    "weight_axle1":   "E74",
    "weight_axle2":   "E75",
    "weight_axle3":   "E76",
    "weight_axle4":   "E77",
    "weight_front":   "E80",
    "weight_rear":    "E81",
    "axle_distance":  "E135",
    "total_length":   "E134",
    "body_length":    "E138",
    "front_axle_to_edge": "E139",
    "rear_axle_to_edge":  "E141",
    "rear_overhang":  "E145",
    # Weighing - bridge
    "bridge_front_axles": "E163",
    "bridge_rear_axles":  "E164",
    "bridge_total":       "E165",
    # Weighing - examiner
    "exam_axle1_right":   "E158",
    "exam_axle1_left":    "E159",
    "exam_axle2_right":   "E161",
    "exam_axle2_left":    "E162",
    # Device 1
    "dev1_name":         "E105",
    "dev1_installer":    "E107",
    "dev1_manufacturer": "E108",
    "dev1_model":        "E109",
    "dev1_serial":       "E110",
    # Device 2
    "dev2_name":         "E112",
    "dev2_installer":    "E114",
    "dev2_manufacturer": "E115",
    "dev2_model":        "E116",
    "dev2_serial":       "E117",
    # Device 3
    "dev3_name":         "E119",
    "dev3_installer":    "E121",
    "dev3_manufacturer": "E122",
    "dev3_model":        "E123",
    "dev3_serial":       "E124",
}


def _cell_val(ws, cell_ref):
    """Read a cell value, return stripped string or empty string."""
    val = ws[cell_ref].value
    if val is None:
        return ""
    return str(val).strip()


def read_secretary_data(excel_path: Path, category: str) -> dict:
    """Read reference data from מזכירה sheet."""
    key = "N" if category in ("N2", "N3") else "M"
    wb = openpyxl.load_workbook(str(excel_path), data_only=True)
    mapping = SECRETARY_CELLS[key]

    # Find the מזכירה sheet
    sheet_name = None
    for sn in wb.sheetnames:
        if "מזכיר" in sn:
            sheet_name = sn
            break
    if not sheet_name:
        wb.close()
        return {}

    ws = wb[sheet_name]
    result = {}
    for field, (_, cell) in mapping.items():
        result[field] = _cell_val(ws, cell)
    wb.close()
    return result


def detect_category(excel_path: Path) -> str:
    """Auto-detect category from מזכירה D17 cell."""
    try:
        wb = openpyxl.load_workbook(str(excel_path), data_only=True)
        for sn in wb.sheetnames:
            if "מזכיר" in sn:
                ws = wb[sn]
                val = _cell_val(ws, "D17")
                wb.close()
                if val:
                    return val
                break
        wb.close()
    except Exception:
        pass
    return "N2"


def write_examiner_data(excel_path: Path, data: dict):
    """Write examiner field data to בוחן sheet."""
    wb = openpyxl.load_workbook(str(excel_path))

    # Find בוחן sheet
    sheet_name = None
    for sn in wb.sheetnames:
        if "בוחן" in sn.strip():
            sheet_name = sn
            break
    if not sheet_name:
        wb.close()
        return False

    ws = wb[sheet_name]
    for field, cell_ref in EXAMINER_CELLS_N.items():
        if field in data and data[field]:
            ws[cell_ref] = data[field]

    wb.save(str(excel_path))
    wb.close()
    return True


def read_deficiencies(excel_path: Path) -> dict:
    """Read deficiency data from פ. ממצאים מסכם sheet."""
    wb = openpyxl.load_workbook(str(excel_path), data_only=True)
    sheet_name = None
    for sn in wb.sheetnames:
        if "ממצאים" in sn:
            sheet_name = sn
            break
    if not sheet_name:
        wb.close()
        return {"pre": [], "post": [], "meta": {}}

    ws = wb[sheet_name]
    meta = {
        "report_num": _cell_val(ws, "A13"),
        "manufacturer": _cell_val(ws, "E13"),
        "license": _cell_val(ws, "H13"),
        "vin": _cell_val(ws, "J13"),
    }

    # Pre-inspection deficiencies (rows 22-27, items 1-6)
    pre = []
    for row in range(22, 28):
        item = {
            "num": row - 21,
            "finding": _cell_val(ws, f"B{row}"),
            "doc_required": _cell_val(ws, f"H{row}"),
            "photo_required": _cell_val(ws, f"I{row}"),
            "reinspect": _cell_val(ws, f"J{row}"),
        }
        pre.append(item)

    # Post-inspection deficiencies (rows 29-34, items 7-12)
    post = []
    for row in range(29, 35):
        item = {
            "num": row - 22,
            "finding": _cell_val(ws, f"B{row}"),
            "doc_required": _cell_val(ws, f"H{row}"),
            "photo_required": _cell_val(ws, f"I{row}"),
            "reinspect": _cell_val(ws, f"J{row}"),
        }
        post.append(item)

    wb.close()
    return {"pre": pre, "post": post, "meta": meta}


def read_examiner_notes(excel_path: Path) -> list:
    """Read examiner deficiency notes from בוחן sheet section 10 (rows 312-319)."""
    wb = openpyxl.load_workbook(str(excel_path), data_only=True)
    sheet_name = None
    for sn in wb.sheetnames:
        if "בוחן" in sn.strip():
            sheet_name = sn
            break
    if not sheet_name:
        wb.close()
        return []

    ws = wb[sheet_name]
    notes = []
    for row in range(312, 320):
        note = {
            "num": f"10.{row - 311}",
            "finding": _cell_val(ws, f"D{row}"),
            "doc_required": _cell_val(ws, f"G{row}"),
            "photo_required": _cell_val(ws, f"H{row}"),
        }
        notes.append(note)
    wb.close()
    return notes


def write_examiner_notes(excel_path: Path, notes: list):
    """Write examiner deficiency notes to בוחן sheet section 10 (rows 312-319)."""
    wb = openpyxl.load_workbook(str(excel_path))
    sheet_name = None
    for sn in wb.sheetnames:
        if "בוחן" in sn.strip():
            sheet_name = sn
            break
    if not sheet_name:
        wb.close()
        return False

    ws = wb[sheet_name]
    for i, note in enumerate(notes[:8]):
        row = 312 + i
        if note.get("finding"):
            ws[f"D{row}"] = note["finding"]
        if note.get("doc_required"):
            ws[f"G{row}"] = note["doc_required"]
        if note.get("photo_required"):
            ws[f"H{row}"] = note["photo_required"]

    wb.save(str(excel_path))
    wb.close()
    return True


def generate_deficiency_pdf(excel_path: Path, manufacturer_name: str) -> bytes:
    """Generate a PDF summarizing deficiencies."""
    deficiencies = read_deficiencies(excel_path)
    examiner_notes = read_examiner_notes(excel_path)

    # Read license + VIN from מזכירה sheet (actual data, not headers)
    license_num = ""
    vin_num = ""
    try:
        wb_meta = openpyxl.load_workbook(str(excel_path), data_only=True)
        for sn in wb_meta.sheetnames:
            if "מזכיר" in sn:
                ws_meta = wb_meta[sn]
                license_num = _cell_val(ws_meta, "D16")
                vin_num = _cell_val(ws_meta, "D25")
                break
        wb_meta.close()
    except Exception:
        pass
    # Fallback: extract from vehicle folder name
    if not license_num or not vin_num:
        vehicle_name = excel_path.stem
        license_num = license_num or vehicle_name
        vin_num = vin_num or ""

    # Use Arial from Windows fonts (supports Hebrew)
    font_path = r"C:\Windows\Fonts\arial.ttf"
    font_bold_path = r"C:\Windows\Fonts\arialbd.ttf"

    pdf = FPDF()
    pdf.add_page()
    pdf.add_font("Arial", "", font_path, uni=True)
    pdf.add_font("Arial", "B", font_bold_path, uni=True)

    # Helper for RTL text
    def rtl(text):
        """Reverse Hebrew text for RTL display in PDF."""
        if not text:
            return ""
        return text[::-1] if any('\u0590' <= c <= '\u05FF' for c in text) else text

    # Title
    pdf.set_font("Arial", "B", 20)
    pdf.cell(0, 14, rtl("חוסרים"), ln=True, align="C")
    pdf.ln(3)

    # Vehicle info
    pdf.set_font("Arial", "", 11)
    pdf.set_text_color(100, 100, 100)
    parts = []
    if manufacturer_name:
        parts.append(manufacturer_name)
    if license_num:
        parts.append(license_num)
    if vin_num:
        parts.append(vin_num)
    info_line = "  |  ".join(parts) if parts else excel_path.stem
    pdf.cell(0, 8, rtl(info_line), ln=True, align="C")
    pdf.set_text_color(0, 0, 0)
    pdf.ln(8)

    # Separator line
    pdf.set_draw_color(200, 200, 200)
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(6)

    # Collect all items: deficiencies + examiner notes in one unified list
    all_items = []
    idx = 1
    for item in deficiencies.get("pre", []):
        if item.get("finding"):
            all_items.append({"num": idx, "text": item["finding"]})
            idx += 1
    for item in deficiencies.get("post", []):
        if item.get("finding"):
            all_items.append({"num": idx, "text": item["finding"]})
            idx += 1
    for note in examiner_notes:
        if note.get("finding") and note["finding"] != "-":
            all_items.append({"num": idx, "text": note["finding"]})
            idx += 1

    if all_items:
        for item in all_items:
            pdf.set_font("Arial", "B", 11)
            pdf.cell(12, 9, str(item["num"]) + ".", border=0, align="C")
            pdf.set_font("Arial", "", 11)
            pdf.cell(0, 9, rtl(item["text"]), border=0, ln=True)
            # Light separator
            pdf.set_draw_color(230, 230, 230)
            pdf.line(22, pdf.get_y(), 200, pdf.get_y())
            pdf.ln(2)
    else:
        pdf.set_font("Arial", "", 12)
        pdf.cell(0, 12, rtl("אין חוסרים"), ln=True, align="C")

    return pdf.output()


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@app.route("/")
def page_manufacturers():
    """Level 1: list manufacturer folders."""
    manufacturers = list_manufacturers()
    return render_template("filelist.html", manufacturers=manufacturers)


@app.route("/dates/<path:manufacturer>")
def page_dates(manufacturer):
    """Level 2: list date folders for a manufacturer."""
    mfr_dir = BASE_DIR / manufacturer
    if not mfr_dir.is_dir():
        return render_template("filelist.html", manufacturers=[],
                               error=f"התיקייה לא קיימת בנתיב: {manufacturer}")
    dates = list_dates(manufacturer)
    return render_template("dates.html", manufacturer=manufacturer, dates=dates)


@app.route("/vehicles/<path:manufacturer>/<date_folder>")
def page_vehicles(manufacturer, date_folder):
    """Level 3: list vehicle folders for a date."""
    date_dir = BASE_DIR / manufacturer / date_folder
    if not date_dir.is_dir():
        return render_template("dates.html", manufacturer=manufacturer, dates=[],
                               error=f"התיקייה לא קיימת בנתיב: {manufacturer}/{date_folder}")
    vehicles = list_vehicles(manufacturer, date_folder)
    return render_template("vehicles.html",
                           manufacturer=manufacturer,
                           date_folder=date_folder,
                           vehicles=vehicles)


@app.route("/category/<path:manufacturer>/<date_folder>/<vehicle>")
def page_category(manufacturer, date_folder, vehicle):
    """Level 4: category selection for a vehicle."""
    excel_path = get_excel_path(manufacturer, date_folder, vehicle)
    if not excel_path.is_file():
        return f"<h2 style='color:red;direction:rtl'>קובץ האקסל לא נמצא: {excel_path.name}</h2>", 404

    return render_template("category.html",
                           manufacturer=manufacturer,
                           date_folder=date_folder,
                           vehicle=vehicle)


@app.route("/inspect/<path:manufacturer>/<date_folder>/<vehicle>/<category>")
def inspect(manufacturer, date_folder, vehicle, category):
    """Main inspection page."""
    excel_path = get_excel_path(manufacturer, date_folder, vehicle)
    photos_dir = get_photos_dir(manufacturer, date_folder, vehicle)

    if not excel_path.is_file():
        return f"<h2 style='color:red;direction:rtl'>קובץ האקסל לא נמצא: {excel_path.name}</h2>", 404
    if not photos_dir.is_dir():
        return f"<h2 style='color:red;direction:rtl'>תיקיית תמונות לא קיימת בנתיב</h2>", 404

    # For categories without forms yet, show placeholder
    if category not in ("N2", "N3"):
        return render_template("inspect_empty.html",
                               category=category,
                               vehicle_name=vehicle,
                               manufacturer=manufacturer,
                               date_folder=date_folder)

    secretary = read_secretary_data(excel_path, category)
    license_num = secretary.get("license", "---")

    # Build a vehicle_path key for the JS (for localStorage + API calls)
    vehicle_key = f"{manufacturer}/{date_folder}/{vehicle}"

    # Read classification options for E87 dropdown
    classifications = read_classification_options(excel_path)

    return render_template("inspect.html",
                           category=category,
                           secretary=json.dumps(secretary, ensure_ascii=False),
                           vehicle_name=vehicle,
                           license_num=license_num,
                           vehicle_key=vehicle_key,
                           manufacturer=manufacturer,
                           date_folder=date_folder,
                           classifications=json.dumps(classifications, ensure_ascii=False))


def read_classification_options(excel_path: Path) -> list:
    """Read T_13 classification dropdown values from גיליון עזר sheet."""
    try:
        wb = openpyxl.load_workbook(str(excel_path), data_only=True)
        ws = None
        for sn in wb.sheetnames:
            if "עזר" in sn:
                ws = wb[sn]
                break
        if not ws:
            wb.close()
            return []

        options = []
        # T_13 values are in column D, starting around row 100
        found_header = False
        for row in range(1, 300):
            val = ws.cell(row=row, column=4).value  # column D
            if val and "T_13" in str(val):
                found_header = True
                continue
            if found_header and val:
                s = str(val).strip()
                if s and s != "-":
                    options.append(s)
            elif found_header and not val:
                # Empty cell after values - check if we've collected enough
                if len(options) > 3:
                    break
        wb.close()
        return options
    except Exception:
        return []


@app.route("/api/classifications")
def api_classifications():
    """Return classification options for E87 dropdown."""
    manufacturer = request.args.get("manufacturer", "")
    date_folder = request.args.get("date", "")
    vehicle = request.args.get("vehicle", "")

    excel_path = get_excel_path(manufacturer, date_folder, vehicle)
    if not excel_path.is_file():
        return jsonify([]), 404

    options = read_classification_options(excel_path)
    return jsonify(options)


@app.route("/api/save_classification", methods=["POST"])
def api_save_classification():
    """Save selected classification to E87 in בוחן sheet."""
    payload = request.get_json()
    manufacturer = payload.get("manufacturer", "")
    date_folder = payload.get("date", "")
    vehicle = payload.get("vehicle", "")
    classification = payload.get("classification", "")

    excel_path = get_excel_path(manufacturer, date_folder, vehicle)
    if not excel_path.is_file():
        return jsonify({"ok": False, "error": "קובץ לא נמצא"}), 404

    try:
        wb = openpyxl.load_workbook(str(excel_path))
        sheet_name = None
        for sn in wb.sheetnames:
            if "בוחן" in sn.strip():
                sheet_name = sn
                break
        if not sheet_name:
            wb.close()
            return jsonify({"ok": False, "error": "גיליון בוחן לא נמצא"}), 404

        ws = wb[sheet_name]
        ws["E87"] = classification
        ws["E88"] = classification
        wb.save(str(excel_path))
        wb.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/secretary")
def api_secretary():
    """Return secretary reference data as JSON."""
    manufacturer = request.args.get("manufacturer", "")
    date_folder = request.args.get("date", "")
    vehicle = request.args.get("vehicle", "")
    category = request.args.get("category", "N2")

    excel_path = get_excel_path(manufacturer, date_folder, vehicle)
    if not excel_path.is_file():
        return jsonify({}), 404

    data = read_secretary_data(excel_path, category)
    return jsonify(data)


@app.route("/api/save", methods=["POST"])
def api_save():
    """Save form data to Excel."""
    payload = request.get_json()
    manufacturer = payload.get("manufacturer", "")
    date_folder = payload.get("date", "")
    vehicle = payload.get("vehicle", "")
    form_data = payload.get("form", {})

    excel_path = get_excel_path(manufacturer, date_folder, vehicle)
    if not excel_path.is_file():
        return jsonify({"ok": False, "error": "קובץ האקסל לא נמצא"}), 404

    try:
        write_examiner_data(excel_path, form_data)
    except Exception as e:
        return jsonify({"ok": False, "error": f"Excel error: {e}"}), 500

    return jsonify({"ok": True})


@app.route("/api/save_photo", methods=["POST"])
def api_save_photo():
    """Save a single photo to the vehicle's תמונות folder."""
    payload = request.get_json()
    manufacturer = payload.get("manufacturer", "")
    date_folder = payload.get("date", "")
    vehicle = payload.get("vehicle", "")
    photo_key = payload.get("key", "photo")
    photo_b64 = payload.get("data", "")

    if not photo_b64 or not photo_b64.startswith("data:"):
        return jsonify({"ok": False, "error": "No photo data"}), 400

    photos_dir = get_photos_dir(manufacturer, date_folder, vehicle)
    if not photos_dir.is_dir():
        return jsonify({"ok": False, "error": "תיקיית תמונות לא קיימת בנתיב"}), 404

    header, b64data = photo_b64.split(",", 1)
    ext = "png" if "png" in header else "jpg"
    img_bytes = base64.b64decode(b64data)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    fname = f"{photo_key}_{ts}.{ext}"
    target = photos_dir / fname
    target.write_bytes(img_bytes)

    return jsonify({"ok": True, "file": str(target)})


@app.route("/api/deficiencies")
def api_deficiencies():
    """Return deficiency data from פ. ממצאים מסכם and examiner notes."""
    manufacturer = request.args.get("manufacturer", "")
    date_folder = request.args.get("date", "")
    vehicle = request.args.get("vehicle", "")

    excel_path = get_excel_path(manufacturer, date_folder, vehicle)
    if not excel_path.is_file():
        return jsonify({"error": "קובץ לא נמצא"}), 404

    deficiencies = read_deficiencies(excel_path)
    examiner_notes = read_examiner_notes(excel_path)
    return jsonify({"deficiencies": deficiencies, "examiner_notes": examiner_notes})


@app.route("/api/save_deficiency_notes", methods=["POST"])
def api_save_deficiency_notes():
    """Save examiner deficiency notes to בוחן sheet section 10."""
    payload = request.get_json()
    manufacturer = payload.get("manufacturer", "")
    date_folder = payload.get("date", "")
    vehicle = payload.get("vehicle", "")
    notes = payload.get("notes", [])

    excel_path = get_excel_path(manufacturer, date_folder, vehicle)
    if not excel_path.is_file():
        return jsonify({"ok": False, "error": "קובץ לא נמצא"}), 404

    try:
        write_examiner_notes(excel_path, notes)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/deficiency_text")
def api_deficiency_text():
    """Return deficiency summary as formatted text for WhatsApp."""
    manufacturer_param = request.args.get("manufacturer", "")
    date_folder = request.args.get("date", "")
    vehicle = request.args.get("vehicle", "")

    excel_path = get_excel_path(manufacturer_param, date_folder, vehicle)
    if not excel_path.is_file():
        return jsonify({"error": "קובץ לא נמצא"}), 404

    deficiencies = read_deficiencies(excel_path)
    examiner_notes = read_examiner_notes(excel_path)

    # Read license + VIN from מזכירה
    license_num = ""
    vin_num = ""
    try:
        wb = openpyxl.load_workbook(str(excel_path), data_only=True)
        for sn in wb.sheetnames:
            if "מזכיר" in sn:
                ws = wb[sn]
                license_num = _cell_val(ws, "D16")
                vin_num = _cell_val(ws, "D25")
                break
        wb.close()
    except Exception:
        pass

    # Build text message
    lines = []
    lines.append("📋 *חוסרים ופערים*")
    if license_num:
        lines.append(f"מס׳ רישוי: {license_num}")
    if vin_num:
        lines.append(f"מס׳ שלדה: {vin_num}")
    lines.append("")

    idx = 1
    has_items = False

    # Pre-inspection
    pre_items = [d for d in deficiencies.get("pre", []) if d.get("finding")]
    if pre_items:
        has_items = True
        lines.append("⚠️ *פערים טרם הבדיקה:*")
        for item in pre_items:
            lines.append(f"{idx}. {item['finding']}")
            idx += 1
        lines.append("")

    # Post-inspection
    post_items = [d for d in deficiencies.get("post", []) if d.get("finding")]
    if post_items:
        has_items = True
        lines.append("🔴 *פערים לאחר בדיקה:*")
        for item in post_items:
            lines.append(f"{idx}. {item['finding']}")
            idx += 1
        lines.append("")

    # Examiner notes
    note_items = [n for n in examiner_notes if n.get("finding") and n["finding"] != "-"]
    if note_items:
        has_items = True
        for note in note_items:
            lines.append(f"{idx}. {note['finding']}")
            idx += 1
        lines.append("")

    if not has_items:
        lines.append("אין חוסרים.")

    return jsonify({"text": "\n".join(lines)})


@app.route("/api/deficiency_pdf")
def api_deficiency_pdf():
    """Generate PDF, save to vehicle folder, and return it."""
    manufacturer = request.args.get("manufacturer", "")
    date_folder = request.args.get("date", "")
    vehicle = request.args.get("vehicle", "")

    excel_path = get_excel_path(manufacturer, date_folder, vehicle)
    if not excel_path.is_file():
        return jsonify({"error": "קובץ לא נמצא"}), 404

    try:
        pdf_bytes = generate_deficiency_pdf(excel_path, manufacturer)

        # Save PDF to vehicle folder with name matching Excel + חוסרים
        vehicle_dir = get_vehicle_path(manufacturer, date_folder, vehicle)
        pdf_filename = f"{vehicle} - חוסרים.pdf"
        pdf_path = vehicle_dir / pdf_filename
        try:
            pdf_path.write_bytes(pdf_bytes)
        except PermissionError:
            # File may be open; save with timestamp suffix
            ts = datetime.now().strftime("%H%M%S")
            pdf_path = vehicle_dir / f"{vehicle} - חוסרים_{ts}.pdf"
            pdf_path.write_bytes(pdf_bytes)

        return send_file(
            io.BytesIO(pdf_bytes),
            mimetype="application/pdf",
            as_attachment=True,
            download_name=pdf_filename
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/deficiency_image")
def api_deficiency_image():
    """Generate a deficiency summary as PNG image for easy WhatsApp sharing."""
    manufacturer_param = request.args.get("manufacturer", "")
    date_folder = request.args.get("date", "")
    vehicle = request.args.get("vehicle", "")

    excel_path = get_excel_path(manufacturer_param, date_folder, vehicle)
    if not excel_path.is_file():
        return jsonify({"error": "קובץ לא נמצא"}), 404

    try:
        from PIL import Image, ImageDraw, ImageFont

        deficiencies = read_deficiencies(excel_path)
        examiner_notes = read_examiner_notes(excel_path)

        # Read license + VIN from מזכירה
        license_num = ""
        vin_num = ""
        try:
            wb_meta = openpyxl.load_workbook(str(excel_path), data_only=True)
            for sn in wb_meta.sheetnames:
                if "מזכיר" in sn:
                    ws_meta = wb_meta[sn]
                    license_num = _cell_val(ws_meta, "D16")
                    vin_num = _cell_val(ws_meta, "D25")
                    break
            wb_meta.close()
        except Exception:
            pass

        # Collect all items
        all_items = []
        for item in deficiencies.get("pre", []):
            if item.get("finding"):
                all_items.append(item["finding"])
        for item in deficiencies.get("post", []):
            if item.get("finding"):
                all_items.append(item["finding"])
        for note in examiner_notes:
            if note.get("finding") and note["finding"] != "-":
                all_items.append(note["finding"])

        # Create image
        W = 1080
        padding = 60
        line_h = 50
        header_h = 200
        content_h = max(len(all_items) * line_h + 40, 100)
        H = header_h + content_h + padding

        img = Image.new("RGB", (W, H), color=(15, 23, 42))
        draw = ImageDraw.Draw(img)

        # Load font
        try:
            font_title = ImageFont.truetype(r"C:\Windows\Fonts\arialbd.ttf", 48)
            font_sub = ImageFont.truetype(r"C:\Windows\Fonts\arial.ttf", 28)
            font_item = ImageFont.truetype(r"C:\Windows\Fonts\arial.ttf", 30)
            font_num = ImageFont.truetype(r"C:\Windows\Fonts\arialbd.ttf", 30)
        except Exception:
            font_title = ImageFont.load_default()
            font_sub = font_title
            font_item = font_title
            font_num = font_title

        y = padding

        # Title
        draw.text((W // 2, y), "חוסרים", fill=(255, 255, 255), font=font_title, anchor="mt")
        y += 70

        # Info line
        info = f"{manufacturer_param}  |  {license_num}  |  {vin_num}"
        draw.text((W // 2, y), info, fill=(148, 163, 184), font=font_sub, anchor="mt")
        y += 50

        # Separator
        draw.line([(padding, y), (W - padding, y)], fill=(51, 65, 85), width=2)
        y += 30

        # Items
        if all_items:
            for i, text in enumerate(all_items):
                num_text = f".{i + 1}"
                # Number on the right
                draw.text((W - padding, y), num_text, fill=(59, 130, 246), font=font_num, anchor="rt")
                # Text
                draw.text((W - padding - 60, y), text, fill=(241, 245, 249), font=font_item, anchor="rt")
                y += line_h
        else:
            draw.text((W // 2, y), "אין חוסרים", fill=(148, 163, 184), font=font_sub, anchor="mt")

        # Save to vehicle folder
        vehicle_dir = get_vehicle_path(manufacturer_param, date_folder, vehicle)
        img_filename = f"{vehicle} - חוסרים.png"
        img_path = vehicle_dir / img_filename
        try:
            img.save(str(img_path))
        except Exception:
            ts = datetime.now().strftime("%H%M%S")
            img_path = vehicle_dir / f"{vehicle} - חוסרים_{ts}.png"
            img.save(str(img_path))

        # Return image
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        buf.seek(0)
        return send_file(buf, mimetype="image/png", as_attachment=True,
                         download_name=img_filename)

    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5555, debug=True)
